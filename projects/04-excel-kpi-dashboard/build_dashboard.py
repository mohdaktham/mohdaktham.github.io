import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

df = pd.read_csv("/tmp/pf/data/superstore.csv")
df.columns = [c.strip() for c in df.columns]
df["Order Date"] = pd.to_datetime(df["Order Date"], errors="coerce")
df = df.dropna(subset=["Order Date"])
df["Month"] = df["Order Date"].dt.to_period("M").astype(str)
df["Year"] = df["Order Date"].dt.year

wb = Workbook()
HDR = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FILL = PatternFill("solid", start_color="1F4E79")
BODY = Font(name="Arial", size=10)
TITLE = Font(name="Arial", bold=True, size=14, color="1F4E79")
thin = Border(bottom=Side(style="thin", color="D9D9D9"))

def write_table(ws, df_t, start_row=1, start_col=1):
    for j, col in enumerate(df_t.columns, start=start_col):
        c = ws.cell(row=start_row, column=j, value=col); c.font = HDR; c.fill = FILL
    for i, row in enumerate(df_t.itertuples(index=False), start=start_row+1):
        for j, v in enumerate(row, start=start_col):
            c = ws.cell(row=i, column=j, value=v); c.font = BODY; c.border = thin

# --- Data sheet (raw monthly aggregates; formulas reference this) ---
ws_d = wb.active; ws_d.title = "Data"
monthly = df.groupby("Month").agg(Sales=("Sales","sum"), Profit=("Profit","sum"), Orders=("Order ID","nunique")).round(0).reset_index()
write_table(ws_d, monthly)
n = len(monthly) + 1
cat = df.groupby("Category").agg(Sales=("Sales","sum"), Profit=("Profit","sum")).round(0).reset_index()
write_table(ws_d, cat, start_row=1, start_col=6)
seg = df.groupby("Segment").agg(Sales=("Sales","sum")).round(0).reset_index()
write_table(ws_d, seg, start_row=1, start_col=10)
reg = df.groupby("Region").agg(Sales=("Sales","sum"), Profit=("Profit","sum")).round(0).reset_index()
reg["Margin"] = None
write_table(ws_d, reg, start_row=8, start_col=6)
for i in range(len(reg)):
    ws_d.cell(row=9+i, column=9, value=f"=H{9+i}/G{9+i}").number_format = "0.0%"
for col, w in [("A",12),("B",12),("C",12),("D",10),("F",14),("G",12),("H",12),("I",10),("J",14),("K",12)]:
    ws_d.column_dimensions[col].width = w
for r in range(2, n+1):
    ws_d.cell(row=r, column=2).number_format = "$#,##0"
    ws_d.cell(row=r, column=3).number_format = "$#,##0"

# --- Dashboard sheet ---
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws["B2"] = "SUPERSTORE SALES KPI DASHBOARD"; ws["B2"].font = TITLE
ws["B3"] = "Source: Superstore retail dataset, 2015-2018 | All KPIs computed with live Excel formulas from the Data sheet"
ws["B3"].font = Font(name="Arial", size=9, italic=True, color="808080")

kpis = [
    ("Total Revenue", f"=SUM(Data!B2:B{n})", "$#,##0"),
    ("Total Profit", f"=SUM(Data!C2:C{n})", "$#,##0"),
    ("Profit Margin", f"=SUM(Data!C2:C{n})/SUM(Data!B2:B{n})", "0.0%"),
    ("Total Orders", f"=SUM(Data!D2:D{n})", "#,##0"),
    ("Avg Order Value", f"=SUM(Data!B2:B{n})/SUM(Data!D2:D{n})", "$#,##0"),
    ("Best Month Rev.", f"=MAX(Data!B2:B{n})", "$#,##0"),
]
for i, (label, formula, fmt) in enumerate(kpis):
    col = 2 + i*2
    lc = ws.cell(row=5, column=col, value=label); lc.font = Font(name="Arial", bold=True, size=9, color="FFFFFF"); lc.fill = FILL; lc.alignment = Alignment(horizontal="center")
    vc = ws.cell(row=6, column=col, value=formula); vc.font = Font(name="Arial", bold=True, size=12); vc.number_format = fmt; vc.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)

lc1 = LineChart(); lc1.title = "Monthly Revenue Trend"; lc1.height = 7.5; lc1.width = 16; lc1.style = 12
lc1.add_data(Reference(ws_d, min_col=2, min_row=1, max_row=n), titles_from_data=True)
lc1.set_categories(Reference(ws_d, min_col=1, min_row=2, max_row=n))
lc1.legend = None; lc1.y_axis.numFmt = "$#,##0"
ws.add_chart(lc1, "B8")

bc = BarChart(); bc.type = "col"; bc.title = "Sales & Profit by Category"; bc.height = 7.5; bc.width = 12; bc.style = 10
bc.add_data(Reference(ws_d, min_col=7, min_row=1, max_col=8, max_row=1+len(cat)), titles_from_data=True)
bc.set_categories(Reference(ws_d, min_col=6, min_row=2, max_row=1+len(cat)))
ws.add_chart(bc, "L8")

pc = PieChart(); pc.title = "Revenue by Segment"; pc.height = 7.5; pc.width = 9
pc.add_data(Reference(ws_d, min_col=11, min_row=1, max_row=1+len(seg)), titles_from_data=True)
pc.set_categories(Reference(ws_d, min_col=10, min_row=2, max_row=1+len(seg)))
ws.add_chart(pc, "B24")

bc2 = BarChart(); bc2.type = "bar"; bc2.title = "Sales by Region"; bc2.height = 7.5; bc2.width = 12; bc2.style = 10
bc2.add_data(Reference(ws_d, min_col=7, min_row=8, max_row=8+len(reg)), titles_from_data=True)
bc2.set_categories(Reference(ws_d, min_col=6, min_row=9, max_row=8+len(reg)))
bc2.legend = None
ws.add_chart(bc2, "J24")

# --- Top Products sheet ---
ws_p = wb.create_sheet("Top Products")
top = (df.groupby(["Category","Sub-Category"]).agg(Sales=("Sales","sum"), Profit=("Profit","sum"), Qty=("Quantity","sum"))
       .round(0).reset_index().sort_values("Sales", ascending=False))
top["Margin"] = None
write_table(ws_p, top)
for i in range(len(top)):
    r = 2 + i
    ws_p.cell(row=r, column=6, value=f"=D{r}/C{r}").number_format = "0.0%"
    ws_p.cell(row=r, column=3).number_format = "$#,##0"
    ws_p.cell(row=r, column=4).number_format = "$#,##0;($#,##0)"
for col, w in [("A",16),("B",16),("C",12),("D",12),("E",8),("F",10)]:
    ws_p.column_dimensions[col].width = w
ws_p.auto_filter.ref = f"A1:F{1+len(top)}"

wb.save("Superstore_KPI_Dashboard.xlsx")
print("saved")
